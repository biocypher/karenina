"""Export benchmarks as reviewer-friendly curation workbooks."""

from __future__ import annotations

import ast
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .benchmark import Benchmark


@dataclass(frozen=True)
class CurationWorkbookSummary:
    """Summary returned after writing a benchmark curation workbook."""

    output_path: Path
    question_count: int
    field_count: int
    draft_count: int
    finished_count: int
    template_parse_error_count: int


@dataclass(frozen=True)
class _TemplateField:
    name: str
    field_type: str
    instruction: str
    reference_value: str
    verification_method: str


_COLUMNS = (
    "Item Number",
    "Source ID",
    "Question",
    "Reference Answer",
    "Functional Area",
    "Subcategory",
    "Complexity",
    "Curation Status",
    "Field Name",
    "Field Type",
    "Judge Instruction",
    "Reference Value",
    "Verification Method",
    "Template Parse Status",
)

_COLUMN_WIDTHS = (12, 14, 54, 34, 22, 28, 12, 18, 34, 20, 72, 28, 28, 28)


def _call_name(node: ast.expr) -> str | None:
    if isinstance(node, ast.Name):
        return node.id
    if isinstance(node, ast.Attribute):
        return node.attr
    return None


def _render_value(node: ast.expr) -> str:
    try:
        value = ast.literal_eval(node)
    except (ValueError, TypeError):
        return ast.unparse(node)
    if isinstance(value, dict | list | tuple | set):
        if isinstance(value, set):
            value = sorted(value, key=str)
        return json.dumps(value, ensure_ascii=False, default=str)
    if value is None:
        return ""
    return str(value)


def _extract_template_fields(template_code: str) -> list[_TemplateField]:
    tree = ast.parse(template_code)
    fields: list[_TemplateField] = []
    for node in ast.walk(tree):
        if not isinstance(node, ast.AnnAssign):
            continue
        if not isinstance(node.target, ast.Name) or not isinstance(node.value, ast.Call):
            continue
        if _call_name(node.value.func) != "VerifiedField":
            continue
        keyword_values = {keyword.arg: keyword.value for keyword in node.value.keywords if keyword.arg is not None}
        fields.append(
            _TemplateField(
                name=node.target.id,
                field_type=ast.unparse(node.annotation),
                instruction=_render_value(keyword_values["description"]) if "description" in keyword_values else "",
                reference_value=_render_value(keyword_values["ground_truth"])
                if "ground_truth" in keyword_values
                else "",
                verification_method=ast.unparse(keyword_values["verify_with"])
                if "verify_with" in keyword_values
                else "",
            )
        )
    return fields


def _metadata_value(metadata: dict[str, Any], *names: str) -> Any:
    for name in names:
        if name in metadata:
            return metadata[name]
    return ""


def _rows_for_benchmark(benchmark: Benchmark) -> tuple[list[list[Any]], Counter[str]]:
    rows: list[list[Any]] = []
    counts: Counter[str] = Counter()
    for item_number, question_id in enumerate(benchmark.get_question_ids(), start=1):
        question = benchmark.get_question(question_id)
        question_metadata = benchmark.get_question_metadata(question_id)
        custom = question_metadata.get("custom_metadata") or {}
        finished = bool(question_metadata["finished"])
        curation_status = "Finished" if finished else "Draft"
        counts[curation_status] += 1

        fields: list[_TemplateField] = []
        parse_status = "No generated template"
        if benchmark.has_template(question_id):
            try:
                fields = _extract_template_fields(benchmark.get_template(question_id))
                parse_status = "Parsed" if fields else "No verified fields found"
            except (SyntaxError, ValueError) as exc:
                counts["parse_errors"] += 1
                parse_status = f"Parse error: {exc}"

        if not fields:
            fields = [_TemplateField("", "", "", "", "")]

        for field in fields:
            rows.append(
                [
                    item_number,
                    _metadata_value(custom, "Source ID", "source_id", "id"),
                    question["question"],
                    question["raw_answer"],
                    _metadata_value(custom, "Area", "Functional Area", "functional_area"),
                    _metadata_value(custom, "Subcategories", "Subcategory", "subcategory"),
                    _metadata_value(custom, "Complexity", "complexity"),
                    curation_status,
                    field.name,
                    field.field_type,
                    field.instruction,
                    field.reference_value,
                    field.verification_method,
                    parse_status,
                ]
            )
            if field.name:
                counts["fields"] += 1
    return rows, counts


def export_curation_workbook(
    benchmark: Benchmark,
    output_path: str | Path,
) -> CurationWorkbookSummary:
    """Write a field-level workbook for benchmark curation and review.

    Template source is parsed with :mod:`ast` and is never executed by this
    exporter.

    Args:
        benchmark: Benchmark to export.
        output_path: Destination ``.xlsx`` path.

    Returns:
        Counts and the resolved workbook path.

    Raises:
        ValueError: If the destination is not an Excel workbook path.
        OSError: If the workbook cannot be written.
    """

    destination = Path(output_path).expanduser().resolve()
    if destination.suffix.lower() != ".xlsx":
        raise ValueError("Curation workbook output must use the .xlsx extension")

    rows, counts = _rows_for_benchmark(benchmark)
    workbook = Workbook()
    review = workbook.active
    review.title = "Curation Review"
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}{len(rows) + 1}"

    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    alternate_fill = PatternFill(fill_type="solid", fgColor="D9E8F5")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for column_number, (label, width) in enumerate(zip(_COLUMNS, _COLUMN_WIDTHS, strict=True), start=1):
        cell = review.cell(row=1, column=column_number, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        review.column_dimensions[get_column_letter(column_number)].width = width

    for row_number, values in enumerate(rows, start=2):
        item_number = int(values[0])
        for column_number, value in enumerate(values, start=1):
            cell = review.cell(row=row_number, column=column_number, value=value)
            cell.alignment = body_alignment
            if item_number % 2 == 0:
                cell.fill = alternate_fill

    summary = workbook.create_sheet("Summary")
    summary_rows = (
        ("Benchmark Name", benchmark.name),
        ("Questions", len(benchmark.get_question_ids())),
        ("Verified Fields", counts["fields"]),
        ("Draft Items", counts["Draft"]),
        ("Finished Items", counts["Finished"]),
        ("Template Parse Errors", counts["parse_errors"]),
    )
    for row_number, (label, value) in enumerate(summary_rows, start=1):
        summary.cell(row=row_number, column=1, value=label).font = Font(bold=True)
        summary.cell(row=row_number, column=2, value=value)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 42

    destination.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(destination)
    except OSError as exc:
        raise OSError(f"Unable to write curation workbook: {destination}") from exc

    return CurationWorkbookSummary(
        output_path=destination,
        question_count=len(benchmark.get_question_ids()),
        field_count=counts["fields"],
        draft_count=counts["Draft"],
        finished_count=counts["Finished"],
        template_parse_error_count=counts["parse_errors"],
    )

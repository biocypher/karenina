"""Tests for the public benchmark curation-workbook exporter."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from karenina.benchmark import Benchmark, TemplateProgressEvent, export_curation_workbook

TEMPLATE = """from karenina.schemas.entities import BaseAnswer, BooleanMatch, VerifiedField

class Answer(BaseAnswer):
    identifies_crbn: bool = VerifiedField(
        description="True when CRBN is identified.",
        ground_truth=True,
        verify_with=BooleanMatch(),
    )
"""


@pytest.mark.unit
class TestCurationWorkbook:
    """Tests for workbook contents and safety behavior."""

    def test_export_uses_final_labels_and_preserves_draft_metadata(self, tmp_path: Path) -> None:
        benchmark = Benchmark.create(name="Open Targets drafts")
        question_id = benchmark.add_question(
            "Which protein is associated with ALS?",
            "CRBN",
            finished=False,
            custom_metadata={
                "id": "001",
                "Area": "Evidence",
                "Subcategories": "Advanced filters",
                "Complexity": 1.0,
            },
        )
        benchmark.add_answer_template(question_id, TEMPLATE)

        summary = export_curation_workbook(benchmark, tmp_path / "review.xlsx")

        assert summary.question_count == 1
        assert summary.field_count == 1
        assert summary.draft_count == 1
        workbook = load_workbook(summary.output_path)
        review = workbook["Curation Review"]
        labels = [cell.value for cell in review[1]]
        assert "Reference Answer" in labels
        assert "Reference Value" in labels
        assert "Curation Status" in labels
        assert all("Expected" not in str(label) for label in labels)
        values = [cell.value for cell in review[2]]
        row = dict(zip(labels, values, strict=True))
        assert row["Source ID"] == "001"
        assert row["Functional Area"] == "Evidence"
        assert row["Curation Status"] == "Draft"
        assert row["Field Name"] == "identifies_crbn"
        assert row["Reference Value"] == "True"
        assert row["Verification Method"] == "BooleanMatch()"
        assert review.freeze_panes == "A2"

    def test_export_does_not_execute_template_source(self, tmp_path: Path) -> None:
        benchmark = Benchmark.create(name="unsafe source")
        question_id = benchmark.add_question("Question?", "Answer", finished=False)
        marker = tmp_path / "should_not_exist"
        malicious_template = f"""from karenina.schemas.entities import BaseAnswer, VerifiedField
Path({str(marker)!r}).touch()
class Answer(BaseAnswer):
    value: str = VerifiedField(description="Value", ground_truth="x")
"""
        benchmark._questions_cache[question_id]["answer_template"] = malicious_template

        export_curation_workbook(benchmark, tmp_path / "safe.xlsx")

        assert not marker.exists()

    def test_export_reports_template_syntax_errors(self, tmp_path: Path) -> None:
        benchmark = Benchmark.create(name="broken draft")
        question_id = benchmark.add_question("Question?", "Answer", finished=False)
        benchmark._questions_cache[question_id]["answer_template"] = "class Answer("

        summary = export_curation_workbook(benchmark, tmp_path / "broken.xlsx")

        assert summary.template_parse_error_count == 1
        workbook = load_workbook(summary.output_path)
        review = workbook["Curation Review"]
        labels = [cell.value for cell in review[1]]
        row = dict(zip(labels, [cell.value for cell in review[2]], strict=True))
        assert str(row["Template Parse Status"]).startswith("Parse error:")

    def test_output_requires_xlsx_extension(self, tmp_path: Path) -> None:
        benchmark = Benchmark.create(name="test")

        with pytest.raises(ValueError, match=".xlsx"):
            export_curation_workbook(benchmark, tmp_path / "review.csv")

    def test_progress_event_is_public(self) -> None:
        event = TemplateProgressEvent(
            event="job_started",
            question_id=None,
            processed_count=0,
            total_count=1,
            successful_count=0,
            failed_count=0,
            percentage=0.0,
            error=None,
            template_code=None,
            task_duration=None,
        )

        assert event.event == "job_started"
